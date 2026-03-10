from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, EmailStr
from pathlib import Path
import openpyxl
from datetime import datetime

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_FILE = Path(__file__).parent / "newsletter_subscribers.xlsx"


def get_workbook():
    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Subscribers"
        ws.append(["Email", "Datum"])
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20
    return wb, wb.active


def email_exists(ws, email: str) -> bool:
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[0].lower() == email.lower():
            return True
    return False


class SubscribeRequest(BaseModel):
    email: str


@app.post("/api/newsletter/subscribe")
async def subscribe(data: SubscribeRequest):
    email = data.email.strip()
    wb, ws = get_workbook()

    if email_exists(ws, email):
        return {"success": False, "message": "Diese E-Mail ist bereits eingetragen."}

    ws.append([email, datetime.now().strftime("%d.%m.%Y %H:%M")])
    wb.save(EXCEL_FILE)

    return {"success": True, "message": "Erfolgreich eingetragen! Danke."}
